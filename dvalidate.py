from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# Load the Excel file
input_file = "new_exl.xlsx"
output_file = "new_exl1.xlsx"

wb = load_workbook(input_file)
ws = wb.active

# Get the existing date_style if it exists, or create a new one
# date_style = wb.named_styles.get('date_style')

# if date_style is None:
#     date_style = NamedStyle(name='date_style', number_format='DD-MM-YYYY')
#     wb.add_named_style(date_style)

# Find the existing date_style by iterating through the list
date_style = None
for style_name in wb.named_styles:
    if style_name == 'date_style':
        date_style = wb.named_styles[style_name]  # Access by name in dictionary
        break

# If not found, create a new one
if date_style is None:
    date_style = NamedStyle(name='date_style', number_format='DD-MM-YYYY')
    wb.add_named_style(date_style)
  
# Get the last row in the sheet
last_row = ws.max_row

# Apply date format to cells in column F
for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=6, max_col=6):
    for cell in row:
        cell.style = date_style

# Validate and correct dates in column F
for row in range(2, last_row + 1):
    cell_f = ws[f'F{row}']
    try:
        date_parts = cell_f.value.split('-')
        if len(date_parts) == 3:
            day, month, year = map(int, date_parts)
            corrected_date = f"{day:02d}-{month:02d}-{year:04d}"
            ws[f'F{row}'] = corrected_date
        else:
            raise ValueError
    except (AttributeError, ValueError):
        ws[f'F{row}'] = None

# Save the changes to the new Excel file
wb.save(output_file)

print("Invalid dates and date formats corrected and saved to the new Excel file successfully.")
