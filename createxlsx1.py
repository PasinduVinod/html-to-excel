import json
from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter  # Importing get_column_letter function
from openpyxl.worksheet.datavalidation import DataValidation

# from openpyxl.utils import get_column_letter

# Read JSON data from file
with open('remapped_table_data.json', 'r') as json_file:
  data = json.load(json_file)

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Write header row
header = [
    'Order', 'Order Line', 'Quantity', 'Edit Quantity', 'Date', 'Edit Date'
]
ws.append(header)

# Set font and background color for header cells
for cell in ws[1]:
  cell.font = Font(color="FFFFFF")  # Set font color to white
  cell.fill = PatternFill(start_color="000000",
                          end_color="000000",
                          fill_type="solid")  # Set background color to black

# Write data rows
for row_data in data:
  row = [
      row_data.get('Order', ''),
      row_data.get('Order Line', ''),
      row_data.get('Quantity', ''),
      row_data.get('Quantity', ''),
      row_data.get('Date', ''),
      row_data.get('Date', '')
  ]
  ws.append(row)

# Lock all cells in the first row
for cell in ws[1]:
  cell.protection = Protection(locked=True)

last_row = len(data) + 1  # Add 1 to account for the header row
print("last row: ", last_row)

# Unlock cells in column D starting from the second row
for row in ws.iter_rows(min_row=2, max_row=last_row+1, min_col=4, max_col=4):
  for cell in row:
    cell.protection = Protection(locked=False)
    cell.font = Font(color="000000")  # Set font color to black
    cell.fill = PatternFill(start_color="00ff00",
                            end_color="00ff00",
                            fill_type="solid")  # Set background color to green

# Apply date format to cells in column D starting from the second row
number_style = NamedStyle(name='number_style', number_format='0')
for cell in ws.iter_rows(min_row=2, max_row=last_row+1, min_col=4, max_col=4):
  for cell in row:
    cell.style = number_style

# Unlock cells in column F starting from the second row
for row in ws.iter_rows(min_row=2, max_row=last_row+1, min_col=6, max_col=6):
  for cell in row:
    cell.protection = Protection(locked=False)
    cell.font = Font(color="000000")  # Set font color to black
    cell.fill = PatternFill(start_color="00ff00",
                            end_color="00ff00",
                            fill_type="solid")  # Set background color to green

# Apply date format to cells in column F starting from the second row
date_style = NamedStyle(name='date_style', number_format='DD-MM-YYYY')
for cell in ws.iter_rows(min_row=2, max_row=last_row+1, min_col=6, max_col=6):
  for cell in row:
    cell.style = date_style

# Set column width to fit content
for column in ws.columns:
  max_length = 0
  for cell in column:
    try:
      if len(str(cell.value)) > max_length:
        max_length = len(str(cell.value))
    except:
      pass
  adjusted_width = (
      max_length +
      2) * 1.2  # Adjusting width for padding and better visibility
  column_letter = get_column_letter(
      column[0].column)  # Convert column index to letter
  ws.column_dimensions[column_letter].width = adjusted_width

# Protect the worksheet to enforce read-only for the specified columns
ws.protection.sheet = True

# Save workbook to Excel file
wb.save('exl.xlsx')

print("Excel file 'exl.xlsx' has been successfully created.")