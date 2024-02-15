from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Load the Excel file
file_path = "exl.xlsx"
output_file = "new_exl.xlsx"

wb = load_workbook(file_path)
ws = wb.active

# Get the last row in the sheet
last_row = ws.max_row

# Add data validation to column D (Edit Quantity)
dv = DataValidation(type="decimal", operator="lessThanOrEqual", formula1="C2", showErrorMessage=True)
dv.error = "Edit Quantity cannot exceed Quantity"
dv.errorTitle = "Invalid Quantity"
dv.prompt = "Please enter a value less than or equal to Quantity in column C"
dv.promptTitle = "Edit Quantity Validation"
dv_range = f"D2:D{last_row}"  # Range for Edit Quantity column
ws.add_data_validation(dv)
dv.add(dv_range)

# Save the changes to the Excel file
wb.save(output_file)

print("Data validation added successfully.")
